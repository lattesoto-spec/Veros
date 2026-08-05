from __future__ import annotations

import random
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parent
RNG = random.Random(260805)
FACILITY = "Banksia Grove Residential Care"
PERIOD_START = date(2026, 7, 1)
PERIOD_END = date(2026, 7, 31)
DATES = [PERIOD_START + timedelta(days=i) for i in range(31)]

NAVY = "19324D"
TEAL = "167D86"
PALE_TEAL = "DDEFF0"
PALE_BLUE = "EAF0F6"
PALE_GOLD = "FFF4D6"
PALE_RED = "FBE4E6"
WHITE = "FFFFFF"
INK = "243341"
GRID = "D7DEE5"
THIN = Side(style="thin", color=GRID)


FIRST_NAMES = [
    "Amelia", "Noah", "Olivia", "Jack", "Isla", "William", "Mia", "Henry",
    "Charlotte", "Thomas", "Grace", "James", "Sophie", "Leo", "Ruby", "Lucas",
    "Chloe", "Ethan", "Zoe", "Samuel", "Matilda", "Alexander", "Ella", "Daniel",
    "Evie", "Liam", "Ava", "Benjamin", "Lucy", "Oliver", "Hannah", "Max",
    "Emily", "Joshua", "Georgia", "Oscar", "Lily", "Sebastian", "Claire", "Isaac",
    "Margaret", "Arthur", "Helen", "George", "Doreen", "Ronald", "Judith", "Peter",
    "Irene", "Kenneth", "Janet", "Graham", "Patricia", "Raymond", "Sylvia", "Colin",
    "Maureen", "Frank", "Valerie", "Barry", "Nora", "Albert", "Jean", "Keith",
]
LAST_NAMES = [
    "Nguyen", "Martin", "Thompson", "Brown", "Wilson", "Taylor", "Anderson", "Lee",
    "Harris", "Clark", "Walker", "Wright", "Roberts", "King", "Hall", "Young",
    "Allen", "Scott", "Green", "Baker", "Campbell", "Mitchell", "Carter", "Evans",
    "Collins", "Murphy", "Kelly", "Singh", "Patel", "Williams", "Jones", "Lewis",
    "Robinson", "Edwards", "Stewart", "Morris", "Cooper", "Cook", "Morgan", "Bell",
    "Reed", "Bailey", "Rivera", "Foster", "Ward", "Gray", "James", "Watson",
    "Brooks", "Wood", "Bennett", "Price", "Hughes", "Russell", "Griffin", "Diaz",
    "Hayes", "Myers", "Ford", "Hamilton", "Graham", "Sullivan", "Wallace", "Woods",
]


def daterange(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def style_title(ws, title: str, subtitle: str, width: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(name="Aptos Display", size=17, bold=True, color=WHITE)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 29
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws.cell(2, 1).font = Font(name="Aptos", size=10, italic=True, color=INK)
    ws.row_dimensions[2].height = 20
    ws.sheet_view.showGridLines = False


def style_table(ws, header_row: int, last_row: int, widths: list[int]):
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[header_row].height = 34
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(widths))}{last_row}"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, max_col=len(widths)):
        for cell in row:
            cell.font = Font(name="Aptos", size=9, color=INK)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F7F9FB")


def style_info_sheet(ws, title: str, subtitle: str, rows: list[tuple[str, object]]):
    style_title(ws, title, subtitle, 5)
    ws["A4"] = "Export field"
    ws["B4"] = "Value"
    for cell in ws[4][:2]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(bold=True, color=WHITE)
    for r, (label, value) in enumerate(rows, start=5):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 1).font = Font(bold=True, color=INK)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.cell(r, 1).border = Border(bottom=THIN)
        ws.cell(r, 2).border = Border(bottom=THIN)
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 62


def set_date_format(ws, cols: list[int], first_row: int, last_row: int):
    for col in cols:
        for row in range(first_row, last_row + 1):
            ws.cell(row, col).number_format = "dd/mm/yyyy"


def set_time_format(ws, cols: list[int], first_row: int, last_row: int):
    for col in cols:
        for row in range(first_row, last_row + 1):
            ws.cell(row, col).number_format = "h:mm AM/PM"


def make_residents():
    residents = []
    for i in range(1, 51):
        rid = f"BG-{1000 + i}"
        first = FIRST_NAMES[40 + (i - 1) % 24]
        last = LAST_NAMES[(i * 5 + 3) % len(LAST_NAMES)]
        admitted = date(2022 + (i % 4), (i * 3) % 12 + 1, (i * 7) % 25 + 1)
        if i in (48, 49, 50):
            admitted = date(2026, 7, 6 + (i - 48) * 8)
        discharged = None
        if i == 45:
            discharged = date(2026, 7, 12)
        elif i == 46:
            discharged = date(2026, 7, 24)
        funding = "AN-ACC Permanent"
        if i in (42, 48):
            funding = "AN-ACC Respite"
        elif i == 47:
            funding = "Transition Care Programme"
        elif i == 49:
            funding = "Private / non-subsidised"
        anacc = "" if funding in ("Transition Care Programme", "Private / non-subsidised") else f"{(i % 13) + 1:02d}-{(i % 4) + 1:02d}"
        residents.append({
            "id": rid,
            "first": first,
            "last": last,
            "name": f"{first} {last}",
            "admitted": admitted,
            "discharged": discharged,
            "funding": funding,
            "anacc": anacc,
            "room": f"{100 + i}{'A' if i % 9 == 0 else ''}",
            "wing": ["Banksia", "Wattle", "Waratah"][(i - 1) % 3],
            "language": ["English", "English", "English", "Italian", "Greek", "Vietnamese"][i % 6],
            "mobility": ["Independent", "1-person assist", "2-person assist", "Hoist transfer"][i % 4],
        })
    return residents


def occupancy_state(resident, day: date):
    if day < resident["admitted"]:
        return None
    if resident["discharged"] and day > resident["discharged"]:
        return None
    rid = resident["id"]
    leave_type = ""
    leave_no = ""
    reason = ""
    included = "Y"
    state = "In facility"
    if rid == "BG-1008" and date(2026, 7, 4) <= day <= date(2026, 7, 10):
        leave_type = "Hospital leave"
        leave_no = (day - date(2026, 7, 4)).days + 1
        state = "Hospital leave – bed held"
    elif rid == "BG-1017" and day >= date(2026, 7, 1):
        leave_type = "Hospital leave"
        leave_no = (day - date(2026, 6, 5)).days + 1
        state = "Hospital leave – bed held" if leave_no <= 28 else "Hospital leave – over 28 days"
        if leave_no > 28:
            included = "N"
            reason = "Hospital leave day 29+"
    elif rid == "BG-1029" and date(2026, 7, 20) <= day <= date(2026, 7, 22):
        leave_type = "Hospital leave"
        leave_no = (day - date(2026, 7, 20)).days + 1
        state = "Hospital leave – bed held"
    elif rid == "BG-1034" and date(2026, 7, 12) <= day <= date(2026, 7, 13):
        leave_type = "Social leave"
        leave_no = (day - date(2026, 7, 12)).days + 1
        state = "Social leave – bed held"
    if resident["funding"] == "Transition Care Programme":
        included, reason = "N", "Transition Care Programme"
    elif resident["funding"] == "Private / non-subsidised":
        included, reason = "N", "Non-AN-ACC private resident"
    return state, included, leave_type, leave_no, reason


def resident_workbook(residents):
    wb = Workbook()
    info = wb.active
    info.title = "Export Information"
    census_rows = []
    for day in DATES:
        for resident in residents:
            state = occupancy_state(resident, day)
            if state is None:
                continue
            status, included, leave_type, leave_no, reason = state
            census_rows.append([
                day, resident["id"], resident["name"], resident["wing"], resident["room"],
                status, resident["funding"], included, leave_type, leave_no,
                resident["anacc"], reason, f"CENSUS-{day:%Y%m%d}-0600",
                datetime.combine(day, time(6, 3)),
            ])
    style_info_sheet(info, "Resident Administration Export", "Synthetic operational sample — not real personal information", [
        ("Facility", FACILITY),
        ("Reporting period", "01/07/2026 to 31/07/2026"),
        ("Generated", "05/08/2026 09:12 AEST"),
        ("Resident master records", len(residents)),
        ("Daily census rows", len(census_rows)),
        ("Source system", "HarbourCare Resident Administration v7.4"),
        ("Export note", "Contains permanent, respite, TCP, private and leave-day examples."),
    ])

    master = wb.create_sheet("Resident Register")
    style_title(master, f"{FACILITY} — Resident Register", "Resident master extract | generated 05 Aug 2026 09:12 AEST", 14)
    master["A3"] = "Filters applied: All residents with activity during July 2026"
    master.merge_cells("A3:N3")
    headers = ["Client No.", "Preferred name", "Family name", "Current Status", "Admission date", "Discharge date", "AN-ACC Funding Class", "Room / Bed", "Care Area", "Funding Programme", "Primary Language", "Mobility Support", "Last Care Review", "System Notes"]
    master.append([])
    master.append(headers)
    for i, resident in enumerate(residents, start=1):
        status = "Discharged" if resident["discharged"] else "Active"
        master.append([
            resident["id"], resident["first"], resident["last"], status,
            resident["admitted"], resident["discharged"], resident["anacc"],
            resident["room"], resident["wing"], resident["funding"],
            resident["language"], resident["mobility"], date(2026, 6, (i % 27) + 1),
            "Funding review pending" if i in (7, 23) else "",
        ])
    style_table(master, 5, master.max_row, [15, 17, 18, 16, 14, 14, 19, 12, 14, 27, 17, 20, 15, 24])
    set_date_format(master, [5, 6, 13], 6, master.max_row)
    master.conditional_formatting.add(f"D6:D{master.max_row}", FormulaRule(formula=['D6="Discharged"'], fill=PatternFill("solid", fgColor=PALE_RED)))

    daily = wb.create_sheet("Daily Occupancy")
    style_title(daily, f"{FACILITY} — Daily Occupancy Ledger", "Midnight-to-midnight census snapshots | July 2026", 14)
    daily["A3"] = "One row per resident per day. ‘Included in care-minute OBD?’ is the source system’s funding exclusion flag."
    daily.merge_cells("A3:N3")
    daily.append([])
    headers = ["Census Date", "Client Ref", "Resident Display Name", "Wing", "Bed Space", "Occupancy Status", "Funding Stream", "Included in care-minute OBD?", "Leave Category", "Consecutive Leave Day", "AN-ACC Class", "Exclusion / Adjustment Reason", "Snapshot Batch", "Snapshot Timestamp"]
    daily.append(headers)
    for row in census_rows:
        daily.append(row)
    style_table(daily, 5, daily.max_row, [14, 14, 23, 13, 12, 27, 27, 18, 18, 17, 15, 28, 24, 21])
    set_date_format(daily, [1], 6, daily.max_row)
    for r in range(6, daily.max_row + 1):
        daily.cell(r, 14).number_format = "dd/mm/yyyy h:mm AM/PM"
    daily.conditional_formatting.add(f"H6:H{daily.max_row}", FormulaRule(formula=['H6="N"'], fill=PatternFill("solid", fgColor=PALE_RED)))

    path = OUT / "resident_census_july_2026.xlsx"
    wb.save(path)
    return path, len(census_rows)


def make_staff():
    staff = []
    specs = [
        ("RN", 8, "Registered Nurse – Level 1", 55.80),
        ("EN", 6, "Enrolled Nurse", 43.60),
        ("PCW", 32, "Personal Care Worker Grade 2", 34.25),
        ("OTHER", 4, "Lifestyle / Administration", 36.10),
    ]
    idx = 0
    for role, count, title, rate in specs:
        for j in range(count):
            idx += 1
            staff.append({
                "id": f"E{5000 + idx}",
                "name": f"{FIRST_NAMES[idx - 1]} {LAST_NAMES[(idx * 7) % len(LAST_NAMES)]}",
                "role": role,
                "title": title if not (role == "PCW" and j % 9 == 0) else "Assistant in Nursing",
                "employment": "Agency" if (role == "RN" and j >= 6) or (role == "PCW" and j >= 29) else ("Permanent Part-Time" if j % 3 else "Permanent Full-Time"),
                "rate": rate + (18.0 if (role == "RN" and j >= 6) else 0),
                "registration": f"NMW000{82000 + idx}" if role in ("RN", "EN") else "",
                "registration_expiry": date(2027, (idx % 10) + 1, 28) if role in ("RN", "EN") else None,
            })
    return staff


def staffing_workbook(staff):
    by_role = {role: [s for s in staff if s["role"] == role] for role in ("RN", "EN", "PCW", "OTHER")}
    shifts = []
    shift_id = 0

    def add_shift(day, worker, start, finish, break_hours, pay_code, unit, direct="Yes", status="Approved", note=""):
        nonlocal shift_id
        shift_id += 1
        overnight_minutes = (datetime.combine(day, finish) - datetime.combine(day, start)).seconds // 60
        paid_hours = round(overnight_minutes / 60 - break_hours, 2) if status == "Approved" else 0
        weekend = day.weekday() >= 5
        penalty = 1.5 if weekend else (1.15 if start >= time(14) else 1.0)
        if worker["employment"] == "Agency":
            penalty = 1.0
        cost = round(paid_hours * worker["rate"] * penalty, 2)
        shifts.append([
            f"TS-{day:%y%m%d}-{shift_id:05d}", worker["id"], worker["name"],
            worker["title"], day, start, finish, break_hours, paid_hours,
            pay_code, unit, worker["employment"], direct, status, cost,
            "Banksia Grove", note,
        ])

    for day_idx, day in enumerate(DATES):
        rn = by_role["RN"]
        en = by_role["EN"]
        pcw = by_role["PCW"]
        # Two morning RNs, one medication/clinical RN, one PM RN and one night RN.
        for k, start_finish in enumerate([(time(6, 45), time(15, 15)), (time(7, 0), time(15, 0)), (time(9, 30), time(17, 30)), (time(14, 45), time(22, 45)), (time(22, 30), time(7, 0))]):
            worker = rn[(day_idx + k) % len(rn)]
            add_shift(day, worker, *start_finish, 0.5, "ORD" if worker["employment"] != "Agency" else "AGENCY", "Clinical Care")
        for k, start_finish in enumerate([(time(7, 0), time(15, 0)), (time(14, 45), time(22, 45))]):
            worker = en[(day_idx + k * 2) % len(en)]
            add_shift(day, worker, *start_finish, 0.5, "ORD", "Residential Care")
        shift_patterns = ([(time(7, 0), time(15, 0))] * 8 + [(time(14, 45), time(22, 45))] * 6 + [(time(22, 30), time(7, 0))] * 3)
        for k, start_finish in enumerate(shift_patterns):
            # Two understaffed dates retain an explicit unfilled line and no worked substitute.
            if (day.day, k) in ((11, 7), (18, 13), (27, 5)):
                vacant = {"id": f"VAC-{day:%d}-{k}", "name": "Vacant Shift", "title": "Personal Care Worker", "employment": "Unfilled", "rate": 0}
                add_shift(day, vacant, *start_finish, 0.5, "UNFILLED", "Residential Care", status="Not Worked", note="Agency replacement not secured")
                continue
            worker = pcw[(day_idx * 3 + k) % len(pcw)]
            add_shift(day, worker, *start_finish, 0.5, "AGENCY" if worker["employment"] == "Agency" else "ORD", "Residential Care")
        # Non-care and absence entries commonly present in payroll exports.
        if day.weekday() < 5:
            worker = by_role["OTHER"][day_idx % len(by_role["OTHER"])]
            add_shift(day, worker, time(9, 0), time(17, 0), 0.5, "ORD", "Administration", direct="No", note="Front office / reception")
        if day.day in (3, 9, 16, 23, 30):
            worker = pcw[(day_idx + 19) % len(pcw)]
            add_shift(day, worker, time(0, 0), time(0, 0), 0, "SICK", "Residential Care", direct="No", status="Not Worked", note="Personal leave – no worked hours")

    wb = Workbook()
    info = wb.active
    info.title = "Pay Run Summary"
    approved = [r for r in shifts if r[13] == "Approved"]
    style_info_sheet(info, "Payroll & Timekeeping Export", "Synthetic operational sample — pay period ending 31 July 2026", [
        ("Entity", "Banksia Grove Care Services Pty Ltd"),
        ("Facility", FACILITY),
        ("Pay period", "01/07/2026 to 31/07/2026"),
        ("Generated", "05/08/2026 10:04 AEST"),
        ("Approved worked lines", len(approved)),
        ("Other / non-worked lines", len(shifts) - len(approved)),
        ("Approved paid hours", round(sum(r[8] for r in approved), 2)),
        ("Approved labour cost ex GST", round(sum(r[14] for r in approved), 2)),
        ("Source system", "WorkTrack Payroll Enterprise 12.8"),
    ])
    info["B11"].number_format = "$#,##0.00"

    directory = wb.create_sheet("Employee Directory")
    style_title(directory, f"{FACILITY} — Employee Directory", "Active workers appearing in the July 2026 payroll extract", 12)
    directory["A3"] = "Includes agency workers and non-direct-care positions for source-system completeness."
    directory.merge_cells("A3:L3")
    directory.append([])
    directory.append(["Employee Number", "Employee Display Name", "Position Description", "Internal Role Group", "Engagement Type", "Home Cost Centre", "Base Hourly Rate", "AHPRA Registration", "Registration Expiry", "Employment Status", "Payroll Location", "Manager"])
    for i, worker in enumerate(staff):
        directory.append([
            worker["id"], worker["name"], worker["title"], worker["role"], worker["employment"],
            "Clinical Care" if worker["role"] in ("RN", "EN") else ("Residential Care" if worker["role"] == "PCW" else "Administration"),
            worker["rate"], worker["registration"], worker["registration_expiry"], "Active", "Banksia Grove", f"Manager {1 + i % 4}",
        ])
    style_table(directory, 5, directory.max_row, [17, 24, 29, 18, 23, 20, 16, 20, 18, 18, 17, 15])
    set_date_format(directory, [9], 6, directory.max_row)
    for r in range(6, directory.max_row + 1):
        directory.cell(r, 7).number_format = "$0.00"

    detail = wb.create_sheet("Approved Time Entries")
    style_title(detail, f"{FACILITY} — Time Entry Detail", "Payroll extract | 01 Jul 2026 – 31 Jul 2026 | includes non-worked source rows", 17)
    detail["A3"] = "Filters in source system: All pay codes; all cost centres; approved and exception lines"
    detail.merge_cells("A3:Q3")
    detail.append([])
    detail.append(["Source Record", "Payroll Emp #", "Team Member", "Award / Position Description", "Work Date", "Actual Start", "Actual Finish", "Unpaid Meal Break (hours)", "Paid Hours", "Pay Code", "Cost Centre", "Agency / Employee", "Counts as direct care?", "Approval Status", "Labour Cost ex GST", "Location", "Payroll Comment"])
    for row in shifts:
        detail.append(row)
    style_table(detail, 5, detail.max_row, [20, 15, 23, 29, 14, 14, 14, 17, 13, 13, 20, 21, 18, 17, 18, 18, 32])
    set_date_format(detail, [5], 6, detail.max_row)
    set_time_format(detail, [6, 7], 6, detail.max_row)
    for r in range(6, detail.max_row + 1):
        detail.cell(r, 8).number_format = "0.00"
        detail.cell(r, 9).number_format = "0.00"
        detail.cell(r, 15).number_format = "$#,##0.00"
    detail.conditional_formatting.add(f"N6:N{detail.max_row}", FormulaRule(formula=['N6<>"Approved"'], fill=PatternFill("solid", fgColor=PALE_GOLD)))

    path = OUT / "worked_staffing_july_2026.xlsx"
    wb.save(path)
    return path, len(approved), len(shifts) - len(approved), staff, shifts


def care_workbook(residents, staff, shifts):
    clinical_staff = [s for s in staff if s["role"] in ("RN", "EN")]
    care_staff = [s for s in staff if s["role"] == "PCW"]
    episodes = []
    daily_summary = defaultdict(lambda: {"completed": 0, "minutes": 0, "exceptions": 0})
    encounter = 0
    for day_idx, day in enumerate(DATES):
        active = [r for r in residents if occupancy_state(r, day) and occupancy_state(r, day)[1] == "Y"]
        for resident_idx, resident in enumerate(active):
            services = [
                ("Personal care", "Morning hygiene and dressing", "PCW", 25 + (resident_idx % 4) * 5, time(6, 5 + (resident_idx * 7) % 50)),
                ("Clinical", "Medication administration", "RN/EN", 10 + (resident_idx % 3) * 5, time(8, (resident_idx * 3) % 55)),
                ("Activities of daily living", "Mobility / transfer assistance", "PCW", 15 + (resident_idx % 3) * 5, time(11, (resident_idx * 7) % 55)),
            ]
            if (resident_idx + day_idx) % 5 == 0:
                services.append(("Clinical", "Skin integrity / wound care", "RN/EN", 20, time(14, (resident_idx * 5) % 55)))
            if (resident_idx + day_idx) % 4 == 0:
                services.append(("Personal care", "Evening continence support", "PCW", 15, time(19, (resident_idx * 4) % 55)))
            for service_idx, (stream, task, role_group, minutes, start) in enumerate(services):
                encounter += 1
                worker_pool = clinical_staff if role_group == "RN/EN" else care_staff
                worker = worker_pool[(day_idx * 7 + resident_idx + service_idx) % len(worker_pool)]
                status = "Completed"
                note = "Recorded at point of care"
                actual_minutes = minutes
                if encounter % 173 == 0:
                    status, actual_minutes, note = "Resident declined", 0, "Resident declined; RN notified"
                elif encounter % 257 == 0:
                    status, actual_minutes, note = "Cancelled", 0, "Duplicate task cancelled"
                finish_dt = datetime.combine(day, start) + timedelta(minutes=actual_minutes)
                record = [
                    f"ENC-{day:%y%m%d}-{encounter:06d}", day, resident["id"], resident["name"],
                    resident["wing"], stream, task, worker["id"], worker["name"], worker["title"],
                    start, finish_dt.time(), actual_minutes, status, "Banksia Grove",
                    "Mobile point-of-care", note,
                ]
                episodes.append(record)
                if status == "Completed":
                    daily_summary[day]["completed"] += 1
                    daily_summary[day]["minutes"] += actual_minutes
                else:
                    daily_summary[day]["exceptions"] += 1

    wb = Workbook()
    info = wb.active
    info.title = "Export Information"
    completed = sum(v["completed"] for v in daily_summary.values())
    exceptions = sum(v["exceptions"] for v in daily_summary.values())
    style_info_sheet(info, "Point-of-Care Activity Export", "Synthetic service-delivery evidence — not a staff payroll or roster file", [
        ("Facility", FACILITY),
        ("Reporting period", "01/07/2026 to 31/07/2026"),
        ("Generated", "05/08/2026 10:22 AEST"),
        ("Completed activity rows", completed),
        ("Declined / cancelled rows", exceptions),
        ("Total rows", len(episodes)),
        ("Source system", "CareNotes Mobile 5.2"),
        ("Evidence note", "Resident-level delivered-care evidence for reconciliation; not worked staffing evidence."),
    ])

    summary = wb.create_sheet("Daily Service Summary")
    style_title(summary, f"{FACILITY} — Daily Service Summary", "Point-of-care activity totals | July 2026", 5)
    summary.append([])
    summary.append(["Service Date", "Completed Activities", "Delivered Minutes", "Exceptions", "Average Minutes / Completed Activity"])
    for day in DATES:
        values = daily_summary[day]
        avg = round(values["minutes"] / values["completed"], 1) if values["completed"] else 0
        summary.append([day, values["completed"], values["minutes"], values["exceptions"], avg])
    style_table(summary, 4, summary.max_row, [16, 22, 20, 14, 30])
    set_date_format(summary, [1], 5, summary.max_row)
    for r in range(5, summary.max_row + 1):
        summary.cell(r, 3).number_format = "#,##0"
        summary.cell(r, 5).number_format = "0.0"

    detail = wb.create_sheet("Activity Detail")
    style_title(detail, f"{FACILITY} — Delivered Care Activity", "Mobile documentation extract | completed and exception records included", 17)
    detail["A3"] = "This export records resident-level service activity. It must not replace payroll/timekeeping evidence."
    detail.merge_cells("A3:Q3")
    detail.append([])
    detail.append(["Encounter ID", "Service Date", "Client Code", "Client Name", "Care Area", "Care Stream", "Task / Intervention", "Delivered By ID", "Worker Display Name", "Worker Classification", "Visit Start", "Visit Finish", "Minutes Delivered", "Completion Status", "Facility", "Device / Entry Source", "Free Text Note"])
    for row in episodes:
        detail.append(row)
    style_table(detail, 5, detail.max_row, [21, 14, 15, 23, 14, 25, 31, 17, 23, 29, 14, 14, 17, 19, 18, 23, 32])
    set_date_format(detail, [2], 6, detail.max_row)
    set_time_format(detail, [11, 12], 6, detail.max_row)
    detail.conditional_formatting.add(f"N6:N{detail.max_row}", FormulaRule(formula=['N6<>"Completed"'], fill=PatternFill("solid", fgColor=PALE_GOLD)))

    path = OUT / "care_delivery_july_2026.xlsx"
    wb.save(path)
    return path, completed, exceptions, len(episodes)


def write_readme(counts):
    resident_days, shifts_worked, shifts_other, care_completed, care_exceptions, care_total = counts
    text = f"""# Realistic CareMin import samples

All names, identifiers and records in this directory are fictional. The files model one synthetic Australian residential aged-care home for July 2026 and are safe to use in demonstrations.

## Minimum compliance inputs

1. `resident_census_july_2026.xlsx` — resident master plus a daily occupied-bed-day ledger.
2. `worked_staffing_july_2026.xlsx` — approved time entries plus the employee and Ahpra credential directory. CareMin classifies its shift rows as actual worked automatically.

## Optional reconciliation input

`care_delivery_july_2026.xlsx` contains resident-level delivered-care evidence. It helps compare documented care against worked staffing, but it is not required for the compliance numerator or occupied-bed-day denominator.

The workbooks deliberately look like exports from three unrelated systems. They contain title rows, summary sheets, unfamiliar headers, extra business columns, varied role labels, agency staff, leave records, non-care rows, overnight shifts and excluded resident days. Their headers do not match CareMin's built-in exact presets, so they exercise the format-learning path.

## Approximate scale

- Resident register: 50 residents
- Daily census ledger: {resident_days:,} rows
- Staffing export: {shifts_worked:,} approved worked rows and {shifts_other:,} non-worked/exception rows
- Care activity export: {care_completed:,} completed activities and {care_exceptions:,} exception rows ({care_total:,} total)

These are import test inputs, not expected regulatory results. Review the mapping summary, automatic evidence classification and row warnings after each upload before relying on the figures.
"""
    (OUT / "README.md").write_text(text, encoding="utf-8")


def validate_openpyxl(paths):
    for path in paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        assert wb.sheetnames
        for ws in wb.worksheets:
            assert ws.max_row > 0 and ws.max_column > 0
        wb.close()


def validate_with_caremin(paths):
    from carelog.ingestion.mapping import run_spec, validate_results
    from carelog.ingestion.presets import matching_spec
    from carelog.ingestion.reader import read_upload

    specs = {
        "resident_census_july_2026.xlsx": {"targets": [
            {"kind": "residents", "sheet": "Resident Register", "fields": {
                "resident_id": {"column": "Client No."},
                "name": {"source": "combine", "columns": ["Preferred name", "Family name"]},
                "ancc_class": {"column": "AN-ACC Funding Class"},
                "admitted_date": {"column": "Admission date", "parse": "date"},
                "discharged_date": {"column": "Discharge date", "parse": "date"},
            }},
            {"kind": "resident_days", "sheet": "Daily Occupancy", "fields": {
                "date": {"column": "Census Date", "parse": "date"},
                "resident_id": {"column": "Client Ref"},
                "resident_name": {"column": "Resident Display Name"},
                "occupied": {"column": "Included in care-minute OBD?", "parse": "boolean"},
                "service_type": {"column": "Funding Stream"},
                "leave_type": {"column": "Leave Category"},
                "leave_day_number": {"column": "Consecutive Leave Day", "parse": "number"},
                "ancc_class": {"column": "AN-ACC Class"},
                "exclusion_reason": {"column": "Exclusion / Adjustment Reason"},
            }},
        ]},
        "worked_staffing_july_2026.xlsx": {"targets": [{
            "kind": "shifts", "sheet": "Approved Time Entries",
            "row_filter": {"column": "Approval Status", "include_values": ["approved"]},
            "fields": {
                "staff_id": {"column": "Payroll Emp #"},
                "staff_name": {"column": "Team Member"},
                "role": {"column": "Award / Position Description", "normalize": "role"},
                "date": {"column": "Work Date", "parse": "date"},
                "start_time": {"column": "Actual Start", "parse": "time"},
                "end_time": {"column": "Actual Finish", "parse": "time"},
                "break_minutes": {"column": "Unpaid Meal Break (hours)", "parse": "number", "multiply": 60},
                "is_direct_care": {"column": "Counts as direct care?", "parse": "boolean"},
                "is_agency": {"column": "Agency / Employee", "value_map": {"agency": True, "permanent part-time": False, "permanent full-time": False}},
                "labour_cost": {"column": "Labour Cost ex GST", "parse": "number"},
            },
        }]},
        "care_delivery_july_2026.xlsx": {"targets": [{
            "kind": "care_episodes", "sheet": "Activity Detail",
            "row_filter": {"column": "Completion Status", "include_values": ["completed"]},
            "fields": {
                "date": {"column": "Service Date", "parse": "date"},
                "resident_id": {"column": "Client Code"},
                "resident_name": {"column": "Client Name"},
                "care_type": {"column": "Care Stream"},
                "care_category": {"column": "Task / Intervention"},
                "staff_id": {"column": "Delivered By ID"},
                "staff_name": {"column": "Worker Display Name"},
                "role": {"column": "Worker Classification", "normalize": "role"},
                "start_time": {"column": "Visit Start", "parse": "time"},
                "end_time": {"column": "Visit Finish", "parse": "time"},
                "minutes": {"column": "Minutes Delivered", "parse": "number"},
            },
        }]},
    }
    for path in paths:
        sheets = read_upload(path.name, path.read_bytes())
        assert matching_spec(sheets) is None, f"{path.name} unexpectedly matched an exact preset"
        results = run_spec(specs[path.name], sheets)
        problems = validate_results(results)
        assert not problems, f"{path.name}: {problems}"
        for result in results:
            assert not result.row_errors, f"{path.name}/{result.kind}: {result.row_errors[:3]}"
            print("caremin_validation", path.name, result.kind, len(result.records), "filtered", result.rows_filtered)


def render_previews(paths):
    from PIL import Image, ImageDraw, ImageFont

    preview_dir = Path("/tmp/caremin_sample_previews")
    preview_dir.mkdir(parents=True, exist_ok=True)
    font_path = "/System/Library/Fonts/Supplemental/Arial.ttf"
    bold_path = "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
    font = ImageFont.truetype(font_path, 12)
    bold = ImageFont.truetype(bold_path, 12)
    label_font = ImageFont.truetype(bold_path, 18)

    for path in paths:
        wb = load_workbook(path, data_only=True)
        rendered = []
        for ws in wb.worksheets:
            rows_to_show = min(ws.max_row, 16)
            cols_to_show = min(ws.max_column, 17)
            widths = []
            for col in range(1, cols_to_show + 1):
                excel_width = ws.column_dimensions[get_column_letter(col)].width or 12
                widths.append(max(70, min(190, int(excel_width * 7))))
            total_width = sum(widths) + 2
            row_height = 30
            canvas = Image.new("RGB", (total_width, rows_to_show * row_height + 44), "#FFFFFF")
            draw = ImageDraw.Draw(canvas)
            draw.rectangle((0, 0, total_width, 42), fill="#EEF3F6")
            draw.text((10, 11), ws.title, font=label_font, fill="#19324D")
            y = 44
            for row in range(1, rows_to_show + 1):
                x = 1
                for col in range(1, cols_to_show + 1):
                    cell = ws.cell(row, col)
                    fill = "#FFFFFF"
                    if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb:
                        rgb = str(cell.fill.fgColor.rgb)
                        if len(rgb) >= 6:
                            fill = f"#{rgb[-6:]}"
                    draw.rectangle((x, y, x + widths[col - 1], y + row_height), fill=fill, outline="#D7DEE5")
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > 24:
                        value = value[:22] + "…"
                    text_colour = "#243341"
                    if cell.font and cell.font.color and cell.font.color.type == "rgb" and cell.font.color.rgb:
                        text_colour = f"#{str(cell.font.color.rgb)[-6:]}"
                    draw.text((x + 4, y + 8), value, font=bold if cell.font and cell.font.bold else font, fill=text_colour)
                    x += widths[col - 1]
                y += row_height
            rendered.append(canvas)
        montage_width = max(im.width for im in rendered)
        montage_height = sum(im.height for im in rendered) + 18 * (len(rendered) - 1)
        montage = Image.new("RGB", (montage_width, montage_height), "#CBD5DF")
        y = 0
        for im in rendered:
            montage.paste(im, (0, y))
            y += im.height + 18
        preview_path = preview_dir / f"{path.stem}.png"
        montage.save(preview_path)
        print("preview", preview_path)
        wb.close()


def main():
    residents = make_residents()
    resident_path, resident_days = resident_workbook(residents)
    staffing_path, shifts_worked, shifts_other, staff, shifts = staffing_workbook(make_staff())
    care_path, care_completed, care_exceptions, care_total = care_workbook(residents, staff, shifts)
    paths = [resident_path, staffing_path, care_path]
    validate_openpyxl(paths)
    validate_with_caremin(paths)
    render_previews(paths)
    write_readme((resident_days, shifts_worked, shifts_other, care_completed, care_exceptions, care_total))
    for path in paths:
        print(path.name, path.stat().st_size)
    print("resident_days", resident_days)
    print("worked_shifts", shifts_worked)
    print("other_shift_rows", shifts_other)
    print("completed_care_episodes", care_completed)
    print("care_exceptions", care_exceptions)


if __name__ == "__main__":
    main()
