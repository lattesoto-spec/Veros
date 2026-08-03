"""Read arbitrary uploads — CSV, Excel, JSON, SQL dumps, SQLite — into a
uniform table structure.

Every file, regardless of source system, becomes a list of Sheet objects:
headers (from the detected header row) + rows of stringified cells.
Downstream code (fingerprinting, mapping, the LLM analyzer) only ever sees
this structure, never the raw file.
"""

import csv
import io
from dataclasses import dataclass, field


@dataclass
class Sheet:
    name: str
    headers: list[str]
    rows: list[dict]  # header -> cell value (str)
    header_row_index: int = 0
    notes: list[str] = field(default_factory=list)


class FileReadError(Exception):
    pass


JSON_EXTENSIONS = (".json", ".jsonl", ".ndjson")
SQL_EXTENSIONS = (".sql",)
SQLITE_EXTENSIONS = (".db", ".sqlite", ".sqlite3")
# Everything a zip may contain (nested zips are not unpacked).
UNPACKABLE_EXTENSIONS = (
    (".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".xls")
    + JSON_EXTENSIONS + SQL_EXTENSIONS + SQLITE_EXTENSIONS
)
SUPPORTED_EXTENSIONS = UNPACKABLE_EXTENSIONS + (".zip",)

# SQLite files start with this regardless of extension.
SQLITE_MAGIC = b"SQLite format 3\x00"


def read_upload(filename: str, data: bytes) -> list[Sheet]:
    """Parse an uploaded file into sheets. Raises FileReadError on unusable files."""
    name = (filename or "").lower()
    if name.endswith(".zip"):
        return _read_zip(data)
    if name.endswith((".xlsx", ".xlsm")):
        return _read_excel(data)
    if name.endswith(".xls"):
        return _read_xls(data)

    from . import structured

    # Sniff the magic bytes first: .db files are routinely named anything.
    if data[:16] == SQLITE_MAGIC or name.endswith(SQLITE_EXTENSIONS):
        return structured.read_sqlite(data)
    if name.endswith(JSON_EXTENSIONS):
        return structured.read_json(filename, data)
    if name.endswith(SQL_EXTENSIONS):
        return structured.read_sql(filename, data)
    # Extension-less or oddly named uploads: fall back to content sniffing so
    # a JSON body doesn't get mangled by the CSV parser.
    head = data.lstrip()[:1]
    if head in (b"[", b"{"):
        return structured.read_json(filename, data)
    # Default: treat as delimited text (csv/tsv/txt).
    return [_read_delimited(filename, data)]


def _read_zip(data: bytes) -> list[Sheet]:
    import zipfile

    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as e:
        raise FileReadError("Could not open zip archive.") from e
    sheets = []
    for info in zf.infolist():
        base = info.filename.rsplit("/", 1)[-1]
        if info.is_dir() or base.startswith((".", "~")):
            continue
        if not base.lower().endswith(UNPACKABLE_EXTENSIONS):  # no nested zips
            continue
        for s in read_upload(base, zf.read(info)):
            s.name = f"{base}:{s.name}" if not s.name.startswith(base) else s.name
            sheets.append(s)
    if not sheets:
        raise FileReadError("Zip archive contains no supported data files.")
    return sheets


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise FileReadError("Could not decode file as text.")


def _read_delimited(filename: str, data: bytes) -> Sheet:
    text = _decode(data)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    raw_rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    raw_rows = [r for r in raw_rows if any((c or "").strip() for c in r)]
    if not raw_rows:
        raise FileReadError("File is empty.")
    header_idx = _find_header_row(raw_rows)
    return _build_sheet(filename or "csv", raw_rows, header_idx)


def _read_excel(data: bytes) -> list[Sheet]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise FileReadError("openpyxl is required for Excel files.") from e
    try:
        # Not read_only: we need merged-cell ranges and hidden-row flags.
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise FileReadError(f"Could not open Excel file: {e}") from e

    sheets = []
    for ws in wb.worksheets:
        # Fill merged ranges with the top-left value so no data is lost.
        merged = list(ws.merged_cells.ranges)
        for rng in merged:
            ws.unmerge_cells(str(rng))
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    ws.cell(row=r, column=c).value = top_left

        hidden = 0
        raw_rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            dim = ws.row_dimensions.get(row_idx)
            if dim is not None and dim.hidden:
                hidden += 1
            cells = [_cell_to_str(c) for c in row]
            if any(c.strip() for c in cells):
                raw_rows.append(cells)
        if not raw_rows:
            continue
        header_idx = _find_header_row(raw_rows)
        sheet = _build_sheet(ws.title, raw_rows, header_idx)
        if merged:
            sheet.notes.append(f"{len(merged)} merged cell range(s) filled with their top-left value")
        if hidden:
            sheet.notes.append(f"{hidden} hidden row(s) included in the import")
        sheets.append(sheet)
    if not sheets:
        raise FileReadError("Workbook contains no data.")
    return sheets


def _read_xls(data: bytes) -> list[Sheet]:
    try:
        import xlrd
    except ImportError as e:
        raise FileReadError("xlrd is required for legacy .xls files.") from e
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        raise FileReadError(f"Could not open .xls file: {e}") from e

    sheets = []
    for ws in book.sheets():
        raw_rows = []
        for i in range(ws.nrows):
            cells = [_xls_cell_to_str(ws.cell(i, j), book.datemode) for j in range(ws.ncols)]
            if any(c.strip() for c in cells):
                raw_rows.append(cells)
        if not raw_rows:
            continue
        header_idx = _find_header_row(raw_rows)
        sheets.append(_build_sheet(ws.name, raw_rows, header_idx))
    if not sheets:
        raise FileReadError("Workbook contains no data.")
    return sheets


def _xls_cell_to_str(cell, datemode) -> str:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        if dt.year < 1902:  # time-only cells sit at the epoch
            return dt.strftime("%H:%M")
        if (dt.hour, dt.minute, dt.second) == (0, 0, 0):
            return dt.strftime("%Y-%m-%d")
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "true" if cell.value else "false"
    return _cell_to_str(cell.value)


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    # datetimes/times/dates stringify sensibly via isoformat
    if hasattr(value, "isoformat"):
        return value.isoformat(sep=" ") if hasattr(value, "hour") and hasattr(value, "year") else value.isoformat()
    return str(value).strip()


def _find_header_row(raw_rows: list[list[str]], scan: int = 10) -> int:
    """Exports often carry title/preamble lines before the real header.
    Pick the row that looks most header-like: many non-empty, mostly
    non-numeric cells, followed by rows of consistent width."""
    best_idx, best_score = 0, -1.0
    for i, row in enumerate(raw_rows[:scan]):
        cells = [(c or "").strip() for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            continue
        non_numeric = sum(1 for c in non_empty if not _looks_numeric(c))
        score = len(non_empty) + 2.0 * (non_numeric / len(non_empty))
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _looks_numeric(s: str) -> bool:
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


def _build_sheet(name: str, raw_rows: list[list[str]], header_idx: int) -> Sheet:
    headers_raw = [(c or "").strip() for c in raw_rows[header_idx]]
    headers, seen = [], {}
    for j, h in enumerate(headers_raw):
        h = h or f"column_{j + 1}"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 1
        headers.append(h)

    rows = []
    for raw in raw_rows[header_idx + 1:]:
        row = {}
        for j, h in enumerate(headers):
            row[h] = (raw[j] or "").strip() if j < len(raw) else ""
        rows.append(row)

    notes = []
    if header_idx > 0:
        preamble = "; ".join(
            " ".join(c for c in r if (c or "").strip()) for r in raw_rows[:header_idx]
        )
        notes.append(f"Skipped {header_idx} preamble row(s): {preamble[:200]}")
    return Sheet(name=name, headers=headers, rows=rows, header_row_index=header_idx, notes=notes)
