"""Report exports: daily CSV, multi-sheet Excel summary, board report PDF."""

import csv
import io
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from carelog.domain.compliance import (
    CALC_VERSION,
    compliance_pct,
    detect_gaps,
    forecast_quarter,
    evidence_summary,
    monthly_breakdown,
    range_breakdown,
    rn_coverage,
    quarterly_tests,
)

DAILY_COLUMNS = [
    ("date", "Date"),
    ("residents", "Active residents"),
    ("total_minutes", "Total care minutes"),
    ("rn_minutes", "RN minutes"),
    ("en_minutes", "EN minutes"),
    ("pca_minutes", "PCA/PCW minutes"),
    ("excluded_minutes", "Withheld eligibility minutes"),
    ("ineligible_minutes", "Ineligible minutes"),
    ("agency_minutes", "Agency minutes"),
    ("care_per_resident", "Care mins / resident"),
    ("rn_per_resident", "RN mins / resident"),
]


def daily_csv(facility, start: date, end: date) -> str:
    rows = range_breakdown(facility.id, start, end)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _, label in DAILY_COLUMNS])
    for r in rows:
        w.writerow([r[key] for key, _ in DAILY_COLUMNS])
    return buf.getvalue()


def summary_xlsx(facility, start: date, end: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Daily"
    ws.append([label for _, label in DAILY_COLUMNS])
    for c in ws[1]:
        c.font = bold
    for r in range_breakdown(facility.id, start, end):
        ws.append([r[key] if key != "date" else r[key].isoformat() for key, _ in DAILY_COLUMNS])

    wm = wb.create_sheet("Monthly")
    wm.append(["Month", "Days with data", "Occupied bed days", "Total care minutes",
               "RN minutes", "EN minutes", "PCA/PCW minutes", "Agency minutes",
               "Care mins / bed day", "RN mins / bed day"])
    for c in wm[1]:
        c.font = bold
    for m in monthly_breakdown(facility.id, start, end):
        wm.append([m["label"], m["days"], m["bed_days"], m["total_minutes"],
                   m["rn_minutes"], m["en_minutes"], m["pca_minutes"], m["agency_minutes"],
                   m["care_per_bed_day"], m["rn_per_bed_day"]])

    wi = wb.create_sheet("Info")
    wi.append(["Facility", facility.name])
    wi.append(["Period", f"{start.isoformat()} to {end.isoformat()}"])
    wi.append(["AN-ACC target (mins/resident/day)", facility.ancc_target])
    wi.append(["RN target (mins/resident/day)", facility.rn_target])
    wi.append(["Calculation version", CALC_VERSION])
    wi.append(["Historical compliance source", "Eligible actual-worked rows only"])

    wt = wb.create_sheet("Compliance tests")
    wt.append(["Test", "Achieved", "Required", "Result", "Shortfall per OBD", "Shortfall total minutes"])
    for c in wt[1]:
        c.font = bold
    tests = quarterly_tests(facility, start, end)
    for row in tests["tests"]:
        wt.append([row["label"], row["achieved"], row["target"],
                   "PASS" if row["passed"] else "FAIL" if row["passed"] is not None else "NO TARGET",
                   row["shortfall"], row["shortfall_minutes"]])

    we = wb.create_sheet("Evidence reconciliation")
    we.append(["Evidence stream", "Minutes / rows", "Treatment"])
    for c in we[1]:
        c.font = bold
    evidence = evidence_summary(facility.id, start, end)
    we.append(["Actual worked", evidence["worked_minutes"], "Included if role eligible"])
    we.append(["Rostered", evidence["rostered_minutes"], "Planning only - excluded from historical compliance"])
    we.append(["Unverified", evidence["unverified_minutes"], "Withheld"])
    we.append(["Delivered care", evidence["delivered_minutes"], "Reconciliation only - never added to staffing hours"])
    we.append(["Resident-day ledger rows", evidence["resident_day_rows"],
               "Denominator source" if evidence["uses_resident_day_ledger"] else "Admission/discharge fallback in use"])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def board_pdf(facility, today: date) -> bytes:
    """One-page board report: headline compliance, monthly table, outlook, risks."""
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=16, spaceAfter=4)
    body = styles["BodyText"]
    small = ParagraphStyle("small", parent=body, fontSize=8, textColor=colors.grey)

    fc = forecast_quarter(facility, today)
    pct = compliance_pct(facility, today)
    cov = rn_coverage(facility.id, today)
    alerts = detect_gaps(facility, today)
    tests = quarterly_tests(facility, fc["q_start"], today) if fc else None
    evidence = evidence_summary(facility.id, fc["q_start"], today) if fc else None

    elements = [
        Paragraph(f"{facility.name} - Board Care Minutes Report", h1),
        Paragraph(today.strftime("Prepared %d %B %Y"), small),
        Spacer(1, 6 * mm),
    ]

    if fc:
        kpi_rows = [
            ["Quarter-to-date average", f"{fc['qtd_avg']:.1f} mins/resident/day",
             f"Target {facility.ancc_target:.0f}"],
            ["Compliance", f"{pct:.1f}% of target" if pct is not None else "Not available", ""],
            ["RN-only quarter-to-date", f"{fc['qtd_rn_avg']:.1f} mins/bed-day",
             f"Required {facility.rn_target * 0.9:.1f}"],
            ["RN + EN quarter-to-date", f"{fc['qtd_rn_en_avg']:.1f} mins/bed-day",
             f"Required {facility.rn_target:.1f}"],
            ["Projected quarter-end average", f"{fc['projected_avg']:.1f}",
             "On track" if fc["on_track"] else "AT RISK"],
            ["RN 24/7 coverage (today)", f"{cov['coverage_pct']:.1f}%", ""],
        ]
        t = Table(kpi_rows, colWidths=[70 * mm, 55 * mm, 40 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ]))
        elements += [t, Spacer(1, 6 * mm)]

        q_start = fc["q_start"]
        months = monthly_breakdown(facility.id, q_start, today)
        if months:
            rows = [["Month", "Bed days", "Care mins/bed day", "RN mins/bed day", "Agency mins"]]
            for m in months:
                rows.append([m["label"], m["bed_days"], f"{m['care_per_bed_day']:.1f}",
                             f"{m['rn_per_bed_day']:.1f}", m["agency_minutes"]])
            t2 = Table(rows, colWidths=[45 * mm, 30 * mm, 40 * mm, 35 * mm, 30 * mm])
            t2.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ]))
            elements += [Paragraph("Quarter by month", styles["Heading3"]), t2, Spacer(1, 6 * mm)]
        if tests:
            test_rows = [["Quarterly test", "Achieved", "Required", "Result"]]
            for row in tests["tests"]:
                test_rows.append([row["label"], row["achieved"], row["target"],
                                  "PASS" if row["passed"] else "FAIL"])
            tt = Table(test_rows, colWidths=[75 * mm, 35 * mm, 35 * mm, 30 * mm])
            tt.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ]))
            elements += [Paragraph("Three quarterly tests", styles["Heading3"]), tt, Spacer(1, 6 * mm)]
        if evidence and (evidence["unverified_minutes"] or not evidence["uses_resident_day_ledger"]):
            elements.append(Paragraph(
                f"Evidence: {evidence['worked_minutes']} worked minutes; "
                f"{evidence['unverified_minutes']} unverified minutes; "
                f"resident-day ledger {'loaded' if evidence['uses_resident_day_ledger'] else 'not loaded'}.",
                body,
            ))
    else:
        elements.append(Paragraph("Insufficient data for quarter analysis.", body))

    elements.append(Paragraph("Risks and alerts", styles["Heading3"]))
    for a in alerts:
        prefix = {"high": "HIGH - ", "medium": "MEDIUM - ", "info": "- "}[a["severity"]]
        elements.append(Paragraph(prefix + a["message"], body))
    elements += [
        Spacer(1, 8 * mm),
        Paragraph(
            f"Generated by CareMin - calculation version {CALC_VERSION} - figures derive from "
            "eligible actual-worked evidence; resolve evidence warnings before external submission.", small),
    ]

    out = io.BytesIO()
    SimpleDocTemplate(out, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm).build(elements)
    return out.getvalue()
